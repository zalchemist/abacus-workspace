#!/usr/bin/env python3
"""
Validacija Excel fajlova prema Excel Serbian Standard pravilima (v3).

Proverava:
1. Konto kolone — da li su tekstualne (vodeće nule očuvane)
2. Širine kolona — minimum 18 za finansijske
3. Excel Table — postojanje, bez duplog AutoFiltera
4. Format brojeva — srpski/evropski standard [$-407]
5. Greške u formulama — #REF!, #VALUE!, #NAME!, #DIV/0!, #N/A
6. Freeze panes — header red zamrznut
7. Formula separator — provera ; vs , u formulama (v3)
8. Total red — provera da nije unutar Table opsega (v3)
9. Datum format — provera DD.MM.YYYY standarda (v3)

Upotreba:
    python validate_excel.py <putanja_do_excel_fajla>
    python validate_excel.py output.xlsx
    python validate_excel.py --verbose output.xlsx
    python validate_excel.py --json output.xlsx
"""

import sys
import os
import re
import json
import argparse

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ GREŠKA: openpyxl nije instaliran. Instaliraj sa: pip install openpyxl")
    sys.exit(1)


# === KONSTANTE ===
SERBIAN_FORMAT_PATTERN = re.compile(r'\[\$-407\]')
MIN_FINANCIAL_WIDTH = 18

# Maksimalan broj redova koji se proveravaju u jednom sheet-u.
# Podesi ovu vrednost prema veličini fajlova koje obrađuješ.
ROW_LIMIT = 1000

KONTO_KEYWORDS = ['konto', 'račun', 'account', 'šifra konta', 'konto broj',
                  'broj računa', 'acc', 'acct']
FORMULA_ERRORS = ['#REF!', '#VALUE!', '#NAME?', '#NAME!', '#DIV/0!', '#N/A',
                  '#NULL!', '#NUM!']
FINANCIAL_KEYWORDS = ['iznos', 'cena', 'ukupno', 'total', 'suma', 'amount',
                      'price', 'cost', 'value', 'saldo', 'duguje', 'potražuje',
                      'količina', 'quantity', 'budget', 'budžet', 'profit',
                      'prihod', 'rashod', 'revenue', 'expense']
DATE_KEYWORDS = ['datum', 'date', 'vreme', 'time', 'od', 'do', 'period',
                 'rok', 'valuta', 'dospeće']
EXPECTED_DATE_FORMAT = 'DD.MM.YYYY'

# Regex za otkrivanje formula sa zarezom kao separatorom argumenata.
# Napomena: openpyxl interno čuva formule sa zarezom (,) kao separator,
# iako Excel sa srpskim/nemačkim locale-om prikazuje tačka-zarez (;).
# Ova provera detektuje slučajeve gde je zarez korišćen u kodu —
# videti sekciju "Formula separator: ; vs ," u dokumentaciji.
FORMULA_COMMA_PATTERN = re.compile(
    r'^='                         # počinje sa =
    r'.*'                         # bilo šta
    r'[A-Z]+\('                   # ime funkcije + otvorena zagrada
    r'[^)]*'                      # sadržaj unutar zagrade
    r','                          # zarez kao separator
    r'[^)]*'                      # ostatak
    r'\)',                         # zatvorena zagrada
    re.IGNORECASE
)

# Pojednostavljen pattern — hvatamo formule koje sadrže tačka-zarez
FORMULA_SEMICOLON_PATTERN = re.compile(
    r'^='
    r'.*'
    r'[A-Z]+\('
    r'[^)]*'
    r';'
    r'[^)]*'
    r'\)',
    re.IGNORECASE
)


class ValidationResult:
    """Rezultat jedne provere."""

    def __init__(self):
        self.passed = []
        self.warnings = []
        self.errors = []

    @property
    def ok(self):
        return len(self.errors) == 0

    def add_pass(self, msg):
        self.passed.append(msg)

    def add_warning(self, msg):
        self.warnings.append(msg)

    def add_error(self, msg):
        self.errors.append(msg)

    def to_dict(self):
        """Konvertuj rezultat u rečnik za JSON izlaz."""
        return {
            'ok': self.ok,
            'passed': self.passed,
            'warnings': self.warnings,
            'errors': self.errors,
            'summary': {
                'passed_count': len(self.passed),
                'warning_count': len(self.warnings),
                'error_count': len(self.errors),
            }
        }


def is_konto_column(column_name):
    """Proveri da li je kolona konto/račun tip."""
    if not column_name:
        return False
    name_lower = str(column_name).lower().strip()
    return any(keyword in name_lower for keyword in KONTO_KEYWORDS)


def is_financial_column(column_name):
    """Proveri da li je kolona finansijska."""
    if not column_name:
        return False
    name_lower = str(column_name).lower().strip()
    return any(keyword in name_lower for keyword in FINANCIAL_KEYWORDS)


def is_date_column(column_name):
    """Proveri da li je kolona datum tip na osnovu naziva."""
    if not column_name:
        return False
    name_lower = str(column_name).lower().strip()
    return any(keyword in name_lower for keyword in DATE_KEYWORDS)


def get_table_ranges(ws):
    """Vrati listu (table_name, min_row, max_row) za sve Table objekte na sheet-u."""
    ranges = []
    tables = ws.tables
    if tables:
        for table_name, table_obj in tables.items():
            # table_obj može biti Table objekat ili string ref zavisno od verzije openpyxl
            ref = table_obj if isinstance(table_obj, str) else table_obj.ref
            try:
                parts = ref.split(':')
                min_row = int(re.search(r'(\d+)', parts[0]).group(1))
                max_row = int(re.search(r'(\d+)', parts[1]).group(1))
                ranges.append((table_name, min_row, max_row))
            except (IndexError, AttributeError):
                pass
    return ranges


def validate_sheet(ws, result, verbose=False):
    """Validiraj jedan sheet."""
    sheet_name = ws.title

    # --- 1. Provera Excel Table-a ---
    tables = ws.tables
    table_ranges = get_table_ranges(ws)
    if tables:
        for table_name, table in tables.items():
            result.add_pass(f"[{sheet_name}] Excel Table pronađen: {table_name}")

            # Proveri dupli AutoFilter
            if ws.auto_filter and ws.auto_filter.ref:
                result.add_error(
                    f"[{sheet_name}] DUPLI AutoFilter detektovan! Table '{table_name}' već ima "
                    f"filter, ali ws.auto_filter.ref = '{ws.auto_filter.ref}'. "
                    f"Ovo izaziva 'repair' dijalog u Excel-u."
                )
            else:
                result.add_pass(f"[{sheet_name}] Nema duplog AutoFiltera")
    else:
        # Nema Table-a — samo info, ne greška
        if ws.auto_filter and ws.auto_filter.ref:
            result.add_pass(f"[{sheet_name}] AutoFilter aktivan (bez Table): {ws.auto_filter.ref}")
        elif verbose:
            result.add_warning(f"[{sheet_name}] Nema Excel Table objekta na sheet-u")

    # --- 2. Detekcija header-a ---
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value)

    if not headers:
        if verbose:
            result.add_warning(f"[{sheet_name}] Nema header reda (red 1 je prazan)")
        return

    # --- 3. Provera konto kolona ---
    for col_idx, header_name in headers.items():
        if is_konto_column(header_name):
            # Proveri da li su vrednosti tekstualne
            all_text = True
            has_leading_zeros = False
            for row in range(2, min(ws.max_row + 1, ROW_LIMIT + 2)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)):
                    all_text = False
                    result.add_error(
                        f"[{sheet_name}] Konto kolona '{header_name}' "
                        f"(kolona {get_column_letter(col_idx)}) "
                        f"sadrži numeričku vrednost u redu {row}: {cell.value}. "
                        f"Konto MORA biti tekst!"
                    )
                    break
                val_str = str(cell.value)
                if val_str and val_str[0] == '0' and len(val_str) > 1:
                    has_leading_zeros = True

            if all_text:
                if has_leading_zeros:
                    result.add_pass(
                        f"[{sheet_name}] Konto kolona '{header_name}' je tekstualna "
                        f"(vodeće nule očuvane ✓)"
                    )
                else:
                    result.add_pass(f"[{sheet_name}] Konto kolona '{header_name}' je tekstualna")

            # Proveri number_format
            for row in range(2, min(ws.max_row + 1, 102)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and cell.number_format != '@' and cell.number_format != 'General':
                    if verbose:
                        result.add_warning(
                            f"[{sheet_name}] Konto kolona '{header_name}', red {row}: "
                            f"format je '{cell.number_format}' umesto '@' (tekst)"
                        )
                    break

    # --- 4. Provera formata finansijskih kolona ---
    for col_idx, header_name in headers.items():
        if is_financial_column(header_name) and not is_konto_column(header_name):
            has_serbian_format = False
            checked = 0
            for row in range(2, min(ws.max_row + 1, 102)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    checked += 1
                    if SERBIAN_FORMAT_PATTERN.search(str(cell.number_format)):
                        has_serbian_format = True
                        break

            if checked > 0:
                if has_serbian_format:
                    result.add_pass(
                        f"[{sheet_name}] Finansijska kolona '{header_name}' ima srpski format"
                    )
                else:
                    sample_cell = ws.cell(row=2, column=col_idx)
                    result.add_error(
                        f"[{sheet_name}] Finansijska kolona '{header_name}' "
                        f"(kolona {get_column_letter(col_idx)}) NEMA srpski format. "
                        f"Trenutni format: '{sample_cell.number_format}'. "
                        f"Očekivan: '[$-407]#,##0.00'"
                    )

    # --- 5. Provera širina kolona ---
    for col_idx, header_name in headers.items():
        if is_financial_column(header_name) and not is_konto_column(header_name):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width and width < MIN_FINANCIAL_WIDTH:
                result.add_error(
                    f"[{sheet_name}] Finansijska kolona '{header_name}' ({col_letter}) "
                    f"ima širinu {width:.1f}, minimum je {MIN_FINANCIAL_WIDTH}. "
                    f"Moguć prikaz '###########'."
                )
            elif width and width >= MIN_FINANCIAL_WIDTH:
                if verbose:
                    result.add_pass(
                        f"[{sheet_name}] Kolona '{header_name}' ({col_letter}) "
                        f"širina: {width:.1f} ✓"
                    )

    # --- 6. Provera formula grešaka ---
    formula_error_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, ROW_LIMIT + 1),
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for error in FORMULA_ERRORS:
                    if error in str(cell.value):
                        formula_error_count += 1
                        result.add_error(
                            f"[{sheet_name}] Formula greška '{error}' pronađena u "
                            f"{cell.coordinate}: {cell.value}"
                        )
                        if formula_error_count >= 10:
                            result.add_error(
                                f"[{sheet_name}] ... više od 10 formula grešaka, prekidam proveru"
                            )
                            return

    if formula_error_count == 0:
        result.add_pass(f"[{sheet_name}] Nema formula grešaka")

    # --- 7. Provera freeze panes ---
    if ws.freeze_panes:
        result.add_pass(f"[{sheet_name}] Freeze panes aktivan: {ws.freeze_panes}")
    elif verbose:
        result.add_warning(f"[{sheet_name}] Freeze panes nije postavljen")

    # --- 8. [v3] Provera formula separatora (; vs ,) ---
    # NAPOMENA: openpyxl interno čuva formule sa zarezom (,) kao separator
    # argumenata, čak i kada Excel sa srpskim locale-om prikazuje tačka-zarez (;).
    # Ova provera detektuje formule sa tačka-zarezom koje su možda ručno
    # postavljene u kodu, što je ispravno za srpski standard ali openpyxl
    # to drugačije interpretira.
    comma_formulas = []
    semicolon_formulas = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, ROW_LIMIT + 1),
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                # Prebroj zareze i tačka-zareze unutar formula (van stringova)
                if ';' in formula:
                    semicolon_formulas.append((cell.coordinate, formula))
                if ',' in formula:
                    # Preskočimo formule koje koriste zarez samo u stringovima
                    # Proveri da li je zarez van navodnika
                    in_string = False
                    has_comma_arg = False
                    for ch in formula:
                        if ch == '"':
                            in_string = not in_string
                        elif ch == ',' and not in_string:
                            has_comma_arg = True
                            break
                    if has_comma_arg:
                        comma_formulas.append((cell.coordinate, formula))

    if semicolon_formulas and not comma_formulas:
        result.add_pass(
            f"[{sheet_name}] Formule koriste tačka-zarez (;) kao separator — "
            f"ispravno za srpski standard"
        )
    elif comma_formulas and not semicolon_formulas:
        # openpyxl interno koristi zarez — ovo je tehnički ispravno
        if verbose:
            result.add_warning(
                f"[{sheet_name}] Formule koriste zarez (,) kao separator "
                f"({len(comma_formulas)} formula). openpyxl interno čuva formule sa "
                f"zarezom, ali Excel sa srpskim locale-om prikazuje tačka-zarez. "
                f"Ovo je tehnički ispravno ali proverite konzistentnost."
            )
        else:
            result.add_pass(
                f"[{sheet_name}] Formule koriste zarez (,) — openpyxl standardni format"
            )
    elif comma_formulas and semicolon_formulas:
        result.add_warning(
            f"[{sheet_name}] Mešovit formula separator! "
            f"{len(semicolon_formulas)} formula sa ';' i {len(comma_formulas)} sa ','. "
            f"Preporučuje se konzistentan stil."
        )
    elif not comma_formulas and not semicolon_formulas:
        if verbose:
            result.add_pass(f"[{sheet_name}] Nema formula sa višestrukim argumentima")

    # --- 9. [v3] Provera da Total redovi nisu unutar Table opsega ---
    total_keywords = ['ukupno', 'total', 'suma', 'zbir', 'grand total', 'subtotal']
    for t_name, t_min, t_max in table_ranges:
        for row_num in range(t_min + 1, min(t_max + 1, ROW_LIMIT + 2)):
            first_cell = None
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col)
                if cell.value is not None:
                    first_cell = cell
                    break
            if first_cell and isinstance(first_cell.value, str):
                val_lower = first_cell.value.lower().strip()
                if any(kw in val_lower for kw in total_keywords):
                    result.add_error(
                        f"[{sheet_name}] Total/Summary red detektovan UNUTAR Table "
                        f"'{t_name}' u redu {row_num}: '{first_cell.value}'. "
                        f"Total red mora biti VAN opsega Table-a "
                        f"(Table opseg: red {t_min}-{t_max})."
                    )

    if table_ranges:
        # Proveri i redove ispod tabele za validni total
        has_total_outside = False
        for t_name, t_min, t_max in table_ranges:
            for row_num in range(t_max + 1, min(t_max + 5, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if any(kw in cell.value.lower().strip() for kw in total_keywords):
                            has_total_outside = True
                            break
                if has_total_outside:
                    break
        if has_total_outside and verbose:
            result.add_pass(f"[{sheet_name}] Total red je pravilno pozicioniran van Table opsega")

    # --- 10. [v3] Provera datum formata (DD.MM.YYYY) ---
    from datetime import datetime as dt
    for col_idx, header_name in headers.items():
        if is_date_column(header_name):
            checked = 0
            has_correct_format = False
            has_wrong_format = False
            for row in range(2, min(ws.max_row + 1, 102)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                checked += 1
                # Proveri da li je datetime objekat sa ispravnim formatom
                if isinstance(cell.value, dt):
                    fmt = cell.number_format
                    if fmt and ('DD.MM.YYYY' in fmt.upper() or 'dd.mm.yyyy' in fmt.lower()):
                        has_correct_format = True
                    elif fmt in ('General', 'mm-dd-yy', 'yyyy-mm-dd', 'm/d/yy',
                                 'd-mmm-yy', 'yyyy-mm-dd h:mm:ss'):
                        has_wrong_format = True
                elif isinstance(cell.value, str):
                    # Proveri da li je string u formatu DD.MM.YYYY
                    if re.match(r'^\d{2}\.\d{2}\.\d{4}$', cell.value.strip()):
                        has_correct_format = True
                    else:
                        has_wrong_format = True

            if checked > 0:
                if has_correct_format and not has_wrong_format:
                    result.add_pass(
                        f"[{sheet_name}] Datum kolona '{header_name}' koristi "
                        f"ispravan format (DD.MM.YYYY)"
                    )
                elif has_wrong_format:
                    sample_cell = ws.cell(row=2, column=col_idx)
                    result.add_warning(
                        f"[{sheet_name}] Datum kolona '{header_name}' "
                        f"(kolona {get_column_letter(col_idx)}) ne koristi "
                        f"standardni srpski format DD.MM.YYYY. "
                        f"Trenutni format: '{sample_cell.number_format}'."
                    )


def validate_workbook(filepath, verbose=False):
    """Validiraj ceo workbook."""
    result = ValidationResult()

    # Proveri da fajl postoji
    if not os.path.exists(filepath):
        result.add_error(f"Fajl ne postoji: {filepath}")
        return result

    # Proveri ekstenziju
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        result.add_warning(
            f"Fajl ima ekstenziju '{ext}'. Validacija radi samo za .xlsx/.xlsm"
        )

    # Učitaj workbook
    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        result.add_error(f"Greška pri otvaranju fajla: {e}")
        return result

    result.add_pass(f"Fajl uspešno otvoren ({len(wb.sheetnames)} sheet-ova)")

    # Proveri svaki sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if verbose:
            print(f"\n  [SHEET: {sheet_name}]")
        validate_sheet(ws, result, verbose)

    wb.close()
    return result


def print_report(filepath, result, verbose=False):
    """Štampaj izveštaj o validaciji."""
    print("\n" + "=" * 60)
    print("  VALIDACIJA EXCEL FAJLA (v3)")
    print("=" * 60)
    print(f"  Fajl: {os.path.basename(filepath)}")
    print(f"  Putanja: {filepath}")
    print(f"  Limit redova: {ROW_LIMIT}")
    print("-" * 60)

    if result.passed:
        print(f"\n  ✅ PROŠLE PROVERE ({len(result.passed)}):")
        for msg in result.passed:
            print(f"     ✅ {msg}")

    if result.warnings:
        print(f"\n  ⚠️  UPOZORENJA ({len(result.warnings)}):")
        for msg in result.warnings:
            print(f"     ⚠️  {msg}")

    if result.errors:
        print(f"\n  ❌ GREŠKE ({len(result.errors)}):")
        for msg in result.errors:
            print(f"     ❌ {msg}")

    print("\n" + "-" * 60)
    if result.ok:
        print("  REZULTAT: ✅ SVE PROVERE PROŠLE")
    else:
        print(f"  REZULTAT: ❌ PRONAĐENO {len(result.errors)} GREŠAKA")
    print("=" * 60 + "\n")

    return 0 if result.ok else 1


def print_json(filepath, result):
    """Štampaj rezultat u JSON formatu za programsku integraciju."""
    output = {
        'file': os.path.basename(filepath),
        'path': os.path.abspath(filepath),
        'row_limit': ROW_LIMIT,
        'version': 'v3',
        **result.to_dict()
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))
    return 0 if result.ok else 1


def main():
    parser = argparse.ArgumentParser(
        description='Validacija Excel fajlova prema Excel Serbian Standard pravilima (v3).'
    )
    parser.add_argument('filepath', help='Putanja do Excel fajla (.xlsx)')
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Prikaži detaljniji izveštaj')
    parser.add_argument('--json', action='store_true', dest='json_output',
                        help='Izlaz u JSON formatu za programsku integraciju')
    args = parser.parse_args()

    result = validate_workbook(args.filepath, args.verbose)

    if args.json_output:
        exit_code = print_json(args.filepath, result)
    else:
        exit_code = print_report(args.filepath, result, args.verbose)

    sys.exit(exit_code)


if __name__ == '__main__':
    main()
